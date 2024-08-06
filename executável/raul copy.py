import customtkinter as tk
from tkinter import filedialog, messagebox
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl
from openpyxl.styles import PatternFill

class AcessoriasBoot:
    def __init__(self, usuario, senha, filepath, root):
        self.usuario = usuario
        self.senha = senha
        self.filepath = filepath
        self.root = root
        self.driver = webdriver.Firefox()
        self.wait = WebDriverWait(self.driver, 10)
        self.obrigacoes_cadastradas = 0  # Inicializa a contagem de obrigações cadastradas

    def login(self):
        driver = self.driver
        driver.get("https://app.acessorias.com")

        # Login
        campo_usuario = self.wait.until(EC.presence_of_element_located((By.NAME, 'mailAC')))
        campo_usuario.clear()
        campo_usuario.send_keys(self.usuario)

        campo_senha = self.wait.until(EC.presence_of_element_located((By.NAME, 'passAC')))
        campo_senha.clear()
        campo_senha.send_keys(self.senha)
        campo_senha.send_keys(Keys.RETURN)
        time.sleep(5)

        # Aguarda o carregamento da página após o login
        self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="M20"]/a'))).click()

        # Loop para iterar sobre as empresas no Excel
        empresas = self.carregar_empresas(self.filepath)

        # Função para obrigação mensal
        linha_atual = 2  # Inicialize a linha atual
        for empresa in empresas:
            nome_obrigacao, departamento, janeiro, fevereiro, marco, abril, maio, junho, julho, agosto, setembro, outubro, novembro, dezembro, prazo_tecnico, competencia, dias_antes = empresa

            # Busca pela obrigação
            novo_obrigacao = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='search']")))
            novo_obrigacao.click()
            novo_obrigacao.send_keys(nome_obrigacao)
            novo_obrigacao.send_keys(Keys.RETURN)
            time.sleep(2)

            try:
                # Verifica se o elemento existe (clcia na busca)
                clicar_obr = driver.find_element(By.XPATH, "//*[@id='divList']/div[2]")
                clicar_obr.click()

                # Seleciona o departamento
                campo_departamento = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrDptID']")))
                campo_departamento.click()
                #time.sleep(3)
                campo_departamento.send_keys(departamento)
                time.sleep(2)  # Pode ser removido se necessário

                # Janeiro 
                campo_janeiro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD01']")))
                campo_janeiro.click()
                campo_janeiro.send_keys(janeiro)
                # time.sleep(2)

                # fevereiro 
                campo_fevereiro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD02']")))
                campo_fevereiro.click()
                campo_fevereiro.send_keys(fevereiro)
                # time.sleep(2)

                # março
                campo_marco = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD03']")))
                campo_marco.click
                campo_marco.send_keys(marco)
                # time.sleep(2)

                # abril
                campo_abril = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD04']")))
                campo_abril.click
                campo_abril.send_keys(abril)
                # time.sleep(2)

                # maio
                campo_maio = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD05']")))
                campo_maio.click
                campo_maio.send_keys(maio)
                # time.sleep(2)

                # junho 
                campo_junho = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD06']")))
                campo_junho.click
                campo_junho.send_keys(junho)
                # time.sleep(2)

                # Julho 
                campo_julho = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD07']")))
                campo_julho.click
                campo_julho.send_keys(julho)
                # time.sleep(2)

                # Agosto
                campo_agosto = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD08']")))
                campo_agosto.click
                campo_agosto.send_keys(agosto)
                # time.sleep(2)

                # Setembro
                campo_setembro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD09']")))
                campo_setembro.click
                campo_setembro.send_keys(setembro)
                # time.sleep(2)

                # Outubro 
                campo_outubro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD10']")))
                campo_outubro.click
                campo_outubro.send_keys(outubro)
                # time.sleep(2)

                # Novembro
                campo_novembro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD11']")))
                campo_novembro.click
                campo_novembro.send_keys(novembro)
                # time.sleep(2)

                # Dezembro
                campo_dezembro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD12']")))
                campo_dezembro.click
                campo_dezembro.send_keys(dezembro)
                # time.sleep(2)

                # Seleciona o prazo técnico
                campo_prazotecnico = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrDAntes']")))
                campo_prazotecnico.click()
                campo_prazotecnico.send_keys(prazo_tecnico)

                # seleciona os dias antes
                dias_antes_obr = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrDAntesTipo']"))
                )
                dias_antes_obr.click()
                dias_antes.send_keys(dias_antes)

                # Definir competência
                defi_comp = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrCompetencia']"))
                )
                defi_comp.click()
                defi_comp.send_keys(competencia)

                # Clica em 'Salvar'
                campo_salvar = driver.find_element(By.XPATH, "//*[@id='main-container']/div[2]/div[2]/div/div/form/div[6]/div[2]/button[1]")
                campo_salvar.click()
                time.sleep(3)  # Pode ser removido se necessário

                # Clica em 'Voltar'
                campo_voltar = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='btBack']")))
                campo_voltar.click()
                time.sleep(3)

            except Exception as e:
                print(f"Erro ao processar a obrigação {nome_obrigacao}: {e}")
                # Se o elemento não existir, cria uma nova obrigação
                self.criar_nova_obrigacao(nome_obrigacao, departamento, janeiro, fevereiro, marco, abril, maio, junho, julho, agosto, setembro, outubro, novembro, dezembro, prazo_tecnico, competencia, dias_antes)

            linha_atual += 1  # Incrementa para a próxima linha

    def criar_nova_obrigacao(self, nome_obrigacao, departamento, janeiro, fevereiro, marco, abril, maio, junho, julho, agosto, setembro, outubro, novembro, dezembro, prazo_tecnico, competencia, dias_antes):
        driver = self.driver
        print(f"Criando nova obrigação: {nome_obrigacao}")
        
        # Implementar a criação de uma nova obrigação aqui
        # Clica no botão para criar nova obrigação
        try:
            # Busca pela obrigação
            novo_obrigacao = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='search']")))
            novo_obrigacao.click()
            novo_obrigacao.send_keys(nome_obrigacao)
            novo_obrigacao.send_keys(Keys.RETURN)
            time.sleep(2)

            try:
                botao_nova_obrigacao = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='main-container']/div[2]/div[2]/div/div/div[1]/form/button[2]")))
                botao_nova_obrigacao.click()
            except: 
                print("Obrigação: ", nome_obrigacao, "Já cadastrado")
                pass

            # Preenche o nome da nova obrigação
            campo_nome = self.wait.until(EC.presence_of_element_located((By.XPATH, "//*[@id='ObrNome']")))
            campo_nome.send_keys(nome_obrigacao)
            
            # Seleciona o departamento
            campo_departamento = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrDptID']")))
            campo_departamento.click()
            #time.sleep(3)
            campo_departamento.send_keys(departamento)
            time.sleep(2)  # Pode ser removido se necessário

            # Janeiro 
            campo_janeiro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD01']")))
            campo_janeiro.click()
            campo_janeiro.send_keys(janeiro)
            # time.sleep(2)

            # fevereiro 
            campo_fevereiro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD02']")))
            campo_fevereiro.click()
            campo_fevereiro.send_keys(fevereiro)
            # time.sleep(2)

            # março
            campo_marco = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD03']")))
            campo_marco.click
            campo_marco.send_keys(marco)
            # time.sleep(2)

            # abril
            campo_abril = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD04']")))
            campo_abril.click
            campo_abril.send_keys(abril)
            # time.sleep(2)

            # maio
            campo_maio = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD05']")))
            campo_maio.click
            campo_maio.send_keys(maio)
            # time.sleep(2)

            # junho 
            campo_junho = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD06']")))
            campo_junho.click
            campo_junho.send_keys(junho)
            # time.sleep(2)

            # Julho 
            campo_julho = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD07']")))
            campo_julho.click
            campo_julho.send_keys(julho)
            # time.sleep(2)

            # Agosto
            campo_agosto = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD08']")))
            campo_agosto.click
            campo_agosto.send_keys(agosto)
            # time.sleep(2)

            # Setembro
            campo_setembro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD09']")))
            campo_setembro.click
            campo_setembro.send_keys(setembro)
            # time.sleep(2)

            # Outubro 
            campo_outubro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD10']")))
            campo_outubro.click
            campo_outubro.send_keys(outubro)
            # time.sleep(2)

            # Novembro
            campo_novembro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD11']")))
            campo_novembro.click
            campo_novembro.send_keys(novembro)
            # time.sleep(2)

            # Dezembro
            campo_dezembro = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrD12']")))
            campo_dezembro.click
            campo_dezembro.send_keys(dezembro)
            # time.sleep(2)

            # Seleciona o prazo técnico
            campo_prazotecnico = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrDAntes']")))
            campo_prazotecnico.click()
            campo_prazotecnico.send_keys(prazo_tecnico)

            # seleciona os dias antes
            dias_antes_obr = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrDAntesTipo']"))
            )
            dias_antes_obr.click()
            dias_antes_obr.send_keys(dias_antes)

            # Definir competência
            defi_comp = self.wait.until(
                EC.element_to_be_clickable((By.XPATH, "//*[@id='ObrCompetencia']"))
            )
            defi_comp.click()
            defi_comp.send_keys(competencia)

            # Clica em 'Salvar'
            campo_salvar = driver.find_element(By.XPATH, "//*[@id='main-container']/div[2]/div[2]/div/div/form/div[6]/div[2]/button[1]")
            campo_salvar.click()
            time.sleep(3)  # Pode ser removido se necessário

            # Clica em 'Voltar'
            campo_voltar = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id='btBack']")))
            campo_voltar.click()
            time.sleep(3)

        except Exception as e:
            print(f"Erro ao criar nova obrigação: {e}")

            # Marca a linha na planilha Excel como cadastrada com sucesso
            self.marcar_obrigacao_cadastrada(self.filepath, linha_atual)
            linha_atual += 1  # Incrementa para a próxima linha

    def carregar_empresas(self, caminho_planilha):
        wb = openpyxl.load_workbook(caminho_planilha)
        ws = wb.active

        empresas = []

        # Suponha que os valores estejam na segunda linha e cada valor esteja em uma coluna separada
        for row in ws.iter_rows(min_row=2, values_only=True):
            nome_obrigacao = row[0]
            departamento = row[1]
            janeiro = row[2]
            fevereiro = row[3]
            marco = row[4]
            abril = row[5]
            maio = row[6]
            junho = row[7]
            julho = row[8]
            agosto = row[9]
            setembro = row[10]
            outubro = row[11]
            novembro = row[12]
            dezembro = row[13]
            prazo_tecnico = row[14]
            dias_antes = row[15]
            competencia = row[16]
            

            empresas.append((
                nome_obrigacao,
                departamento,
                janeiro,
                fevereiro,
                marco,
                abril,
                maio,
                junho,
                julho,
                agosto,
                setembro,
                outubro,
                novembro,
                dezembro,
                prazo_tecnico,
                competencia,
                dias_antes))

        return empresas

    def marcar_obrigacao_cadastrada(self, caminho_planilha, linha):
        wb = openpyxl.load_workbook(caminho_planilha)
        ws = wb.active
        cell = ws.cell(row=linha, column=1)  # A segunda coluna é a coluna 'B'
        fill = PatternFill(start_color="00BF63", end_color="00BF63", fill_type="solid")
        cell.fill = fill
        wb.save(caminho_planilha)

def open_file(root, app):
    filepath = filedialog.askopenfilename(initialdir="/", title="Selecione a planilha", filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*")))
    app.filepath = filepath

class App:
    def __init__(self, root):
        self.root = root
        self.filepath = None
        self.obrigacao_cadastrada_var = tk.StringVar()
        self.obrigacoes_cadastradas_var = tk.StringVar()

        label_user = tk.CTkLabel(root, text="Usuário:")
        label_user.grid(row=0, column=0, padx=10, pady=5)
        self.entry_user = tk.CTkEntry(root)
        self.entry_user.grid(row=0, column=1, padx=10, pady=5)

        label_pass = tk.CTkLabel(root, text="Senha:")
        label_pass.grid(row=1, column=0, padx=10, pady=5)
        self.entry_pass = tk.CTkEntry(root, show="*")
        self.entry_pass.grid(row=1, column=1, padx=10, pady=5)

        btn_open_file = tk.CTkButton(root, text="Selecione a planilha", command=lambda: open_file(self, self))
        btn_open_file.grid(row=2, column=1, columnspan=2, padx=10, pady=5)

        btn_login = tk.CTkButton(root, text="Cadastrar", command=self.login)
        btn_login.grid(row=3, column=1, columnspan=2, padx=10, pady=5)

        label_obrigacoes_cadastradas = tk.CTkLabel(root, text="Obrigações cadastradas:")
        label_obrigacoes_cadastradas.grid(row=5, column=0, padx=10, pady=5)
        label_obrigacoes_cadastradas_count = tk.CTkLabel(root, textvariable=self.obrigacoes_cadastradas_var)
        label_obrigacoes_cadastradas_count.grid(row=5, column=1, padx=10, pady=5)

    def login(self):
        usuario = self.entry_user.get()
        senha = self.entry_pass.get()
        if usuario and senha and self.filepath:
            boot = AcessoriasBoot(usuario, senha, self.filepath, self)
            boot.login()
        else:
            messagebox.showerror("Erro", "Preencha todos os campos e selecione a planilha")

root = tk.CTk()
root.title("Cadastro de Obrigações")
root.geometry("600x250")

app = App(root)

root.mainloop()
